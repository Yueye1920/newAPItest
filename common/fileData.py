# coding=utf-8

import openpyxl

class FileReader:

    @staticmethod
    def readExcelToDict(flie_path='../data/excel/example.xlsx',sheet_name='login'):
        # 打开Excel文件或者创建新文件
        try:
            workbook = openpyxl.load_workbook(flie_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet =workbook.create_sheet(sheet_name)
        # 将数据存储为字典
        data = []

        # 获取列名，从第3行获取列名
        headers = [cell.value for cell in worksheet[2]]


        # 获取数据，从第4行获取数据
        for row in worksheet.iter_rows(min_row=3,values_only=True):
            data.append(dict(zip(headers,row)))
        workbook.close()
        return data

    @staticmethod
    def writeExcelToDict(flie_path='../data/excel/example.xlsx', sheet_name='login',row=None,column=None,value=None):
        # 打开Excel文件或者创建新文件
        try:
            workbook = openpyxl.load_workbook(flie_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        worksheet.cell(row=row,column=column).value=value

        workbook.save(flie_path)



if __name__ == '__main__':
    data=FileReader.readExcelToDict()
    print(data)